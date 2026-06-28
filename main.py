from stravalib import Client
from pathlib import Path
import json
import pandas as pd
from pandas import DatetimeTZDtype
import numpy as np
import matplotlib.pyplot as plt
from matplotlib.backends.backend_pdf import PdfPages
from matplotlib.gridspec import GridSpec
import matplotlib.ticker as ticker
import time
from datetime import date, datetime, timedelta
from io import BytesIO
import pandas as pd
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment
from openpyxl.drawing.image import Image as XLImage    
import argparse
from models import WorkbookSheetConfig, PdfConfig
from strava_config import ACCESS_TOKEN, REFRESH_TOKEN, EXPIRES_AT
    
################################
# CONSTANTS
################################
WEEKLY_MILEAGE_PLOT_TICK_FREQUENCY = 5
FIG_SIZE = (12, 8)
RACE_TYPES = [
    "400m",
    "1/2 mile",
    "1K",
    "1 mile",
    "2 mile",
    "5K",
    "10K",
    "15K",
    "10 mile",
    "20K",
    "Half-Marathon",
    "Marathon"
]
WEEK_FREQ = "W-MON"
DATE_FORMAT = "%Y-W%U"
KM_TO_METERS = 1000
MILE_TO_METERS = 1609.344
DATA_DIR = "strava_data"
ACTIVITIES_FILE = "activities"
PR_FILE = "personal_records"
OUTPUT_DIR = Path("reports")


################################
# ACTIVITIES & DETAILS
################################
def get_strava_client():
    """
    Create a Strava client from the access & refresh tokens.
    """
    client = Client(access_token=ACCESS_TOKEN, refresh_token=REFRESH_TOKEN, token_expires=int(EXPIRES_AT))   
    # print(client.get_athlete())  # Get current athlete details 
    return client


def get_all_activities(start_date: date | None, end_date: date | None) -> dict:
    """
    Get all historical activities for an athlete within the date range.
    Store activities by type:
    {
        "run": [
            {
                "name": "Bucktown 5k",
                "distance": 5040.5,
                "id": 1
                ...
            },
            {
                "name": "Boston 5k",
                "distance": 5040.5,
                "id": 123
                ...
            },
        ],
        "swim": [
            "name": "Afternoon Swim",
            "distance": 1600.2,
            "moving_time": 1917,
            "id": 18521721976,
            ...
        ]
        ...
    }

    Returns:
        dict: dictionary of activities
    """
    print("Loading activities from the Strava API...")
    
    # Fetch all Strava activities
    client = get_strava_client()
    activities = client.get_activities(
        before=datetime(end_date.year, end_date.month, end_date.day), 
        after=datetime(start_date.year, start_date.month, start_date.day)
    )
    
    # Classify activities into the following sport types
    activities_map = {
        "run": [],
        "swim": [],
        "ride": [],
        "other": []
    }
    
    for i, activity in enumerate(activities):
        # Extract select attributes from Stravalib's SummaryActivity model
        to_include = [
            "id", "type", "workout_type", "sport_type", "name", "distance", "elapsed_time", "moving_time", "start_date"
        ]
        activity_dict = json.loads(activity.model_dump_json(include=to_include)) # dump to JSON first to convert non-JSON serializable types (i.e. datetime) to string
        
        match activity.sport_type:
            case "Run":
                activities_map["run"].append(activity_dict)
            case "Swim":
                activities_map["swim"].append(activity_dict)
            case "Ride":
                activities_map["ride"].append(activity_dict)
            case _:
                activities_map["other"].append(activity_dict)

    write_dict_to_json(activities_map, Path(DATA_DIR), 'activities')    
    return activities_map


def get_activity_details(activity_id: str) -> dict:
    """
    Get activity details, like best efforts.
    Note: Each call to client.get_activity() is a single API call.

    Args:
        activity_id (str): _description_

    Returns:
        dict: _description_
    """
    # Get activity details and its best efforts
    client = get_strava_client()
    detailed_activity = client.get_activity(activity_id=activity_id, include_all_efforts=True)
    
    # Extract select attributes from Stravalib's DetailedActivity model
    to_include = [
        "name", "distance", "elapsed_time", "best_efforts"
    ]
    detailed_activity_dict = json.loads(detailed_activity.model_dump_json(include=to_include)) 
    return detailed_activity_dict


################################
# DATA MANIPULATION - computing PRs, creating dataframes underlying plots, calculating summary statistics
################################
def get_personal_records(activities_map: dict, top_N: int = 0) -> dict:
    """
    Given all activities, find the top N running race PRs for all distances in RACE_TYPES.
    Endpoint /activities/{id} returns an activity's "best_efforts" for each distance in RACE_TYPES.
    To limit the number of API requests we'll make, we'll only get details for races (i.e workout_type = 1).

    Args:
        activities_map (dict): activities dictionary
        top_N (int, optional): number of PRs to track for each distance. Defaults to 0.

    Returns:
        dict: dictionary of PRs
    """
    print("Extracting race PRs from activity data...")
    
    pr_map = {}
    
    # Filter thru the "run" activities
    running_activities = activities_map["run"]

    for run_activity in running_activities:
        # Only fetch details for races (i.e. "workout_type" == 1)
        if run_activity["workout_type"] == 1:
            activity_details = get_activity_details(run_activity["id"])

            # Get the activity's best efforts: each activity will contain best efforts for all distances in RACE_TYPES shorter than the activity
            best_efforts = activity_details["best_efforts"]

            for best_effort in best_efforts:
                distance = best_effort["name"]

                # Add activity name & date to the best_effort
                best_effort["activity_name"] = run_activity["name"]
                best_effort["date"] = run_activity["start_date"]

                # Add the best effort to the list of PRs for that distance
                # Then, sort the PRs by elapsed time (fastest first)
                pr_map.setdefault(distance, []).append(best_effort) # adds the distance to the PR Map if it's the first race of its distance type
                pr_map[distance] = sorted(pr_map[distance], key=lambda x: x["elapsed_time"])

                # Filter to the top N PRs for each distance, if specified
                if top_N > 0:
                    pr_map[distance] = pr_map[distance][:top_N]
            
    write_dict_to_json(pr_map, Path(DATA_DIR), 'personal_records')
    return pr_map

def get_activities_and_prs(start_date: date | None, end_date: date | None, top_N: int = 0) -> tuple[dict, dict]:
    """
    Get all activities and the top N race PRs within the specified date range

    Args:
        start_date (date | None): start of the date range
        end_date (date | None): end of the date range
        top_N (int, optional): number of PRs to record for each race distance. Defaults to 0.

    Returns:
        tuple[dict, dict]: dict of activities & dict of race PRs
    """
    activities_map = get_all_activities(start_date, end_date)
    pr_map = get_personal_records(activities_map, top_N)
    return activities_map, pr_map

def validate_pr_map(pr_map: dict, top_N: int) -> bool:
    """
    Check if the PR dictionary contains the top N PRs for each distance.
    
    Args:
        pr_map (dict): PR dictionary
        top_N (int): number of PRs the PR dictionary should contain for each distance

    Returns:
        bool: True/False if the PR dictionary contains sufficient data
    """
    # The PR map contains the top N PRs for each distance if at least one of the distances has at least N data points.
    max_num_races = 0
    for distance, races in pr_map.items():
        max_num_races = max(max_num_races, len(races))
        if max_num_races >= top_N:
            return True

    return False


def create_weekly_mileage_df(running_activities: dict) -> pd.DataFrame:
    """
    Create a dataframe containing total mileage per week.

    Args:
        running_activities (dict): dictionary of running activities

    Returns:
        pd.DataFrame: DataFrame containing weekly mileage
    """
    df = pd.DataFrame(running_activities)

    # Convert 'date' column to datetime & 'distance' to miles
    df['start_date'] = pd.to_datetime(df['start_date'])
    df['distance (miles)'] = df['distance'] / MILE_TO_METERS

    # Group mileage by week: pd.Grouper will create week bins for every week between the start & end dates
    weekly_totals = df.groupby(pd.Grouper(key='start_date', freq=WEEK_FREQ))['distance (miles)'].sum().reset_index()

    # Format dates for better readability in the plot
    weekly_totals['date'] = weekly_totals['start_date'].dt.strftime("%Y-%m-%d")
    weekly_totals['year-week'] = weekly_totals['start_date'].dt.strftime(DATE_FORMAT)
    
    # Reorganize column order to write to the excel workbook
    weekly_totals = weekly_totals.reindex(columns=['date', 'year-week', 'distance (miles)'])

    return weekly_totals


def create_pr_df(personal_records: dict, race: str) -> pd.DataFrame:
    """
    Create a Dataframe containing all PRs for the race distance.
    For each PR, the dataframe will contain the date and the finish time.
    
    Args:
        personal_records (dict): dictionary of PRs
        race (str): the race distance to track

    Returns:
        pd.DataFrame: DataFrame of PRs for the race distance
    """
    # Get PRs for the specified race
    pr_dict = personal_records[race]
    
    # Fetch the date & elapsed time for each PR and store them in a DataFrame
    race_dates = [race["start_date"] for race in pr_dict]
    race_times = [race["elapsed_time"] for race in pr_dict]
    pr_data = {
        "date": race_dates,
        "time (seconds)": race_times
    }
    pr_df = pd.DataFrame(pr_data)
 
    # Convert date column to YYYY-MM-DD format, and create new column for year-week 
    pr_df['date'] = pd.to_datetime(pr_df['date']).dt.strftime("%Y-%m-%d")
    pr_df['year-week'] = pd.to_datetime(pr_df['date']).dt.strftime(DATE_FORMAT)
    
    # Create column to track race time in terms of H:M:S (this will be displayed in the excel spreadsheet)
    pr_df['time (H:M:S)'] = pr_df['time (seconds)'].apply(lambda x: time.strftime('%H:%M:%S', time.gmtime(x)))
    
    # Specify column order to write to excel spreadsheet
    pr_df = pr_df.reindex(columns=['date', 'year-week', 'time (seconds)', 'time (H:M:S)'])

    # Sort PRs in chronological order (required for line plots to track from left to right)
    pr_df.sort_values(by='date', inplace=True)

    return pr_df


def mileage_vs_race_statistics(running_activities: dict) -> pd.DataFrame:
    """
    Compute summary statistics analyzing weekly mileage.
    Eventually, this may include statistics to quantify the relationship between weekly mileage vs. race times.
    
    Args:
        running_activities (dict): dictionary of running activities

    Returns:
        pd.DataFrame: DataFrame containing weekly mileage data
    """
    weekly_mileage_df = create_weekly_mileage_df(running_activities)

    # Computing change in weekly mileage requires multiple weeks of data
    if len(weekly_mileage_df.index) > 1:
        avg_mileage_delta, _ = np.polyfit(weekly_mileage_df.index, weekly_mileage_df['distance (miles)'], 1)
        avg_mileage_delta_str = f"{avg_mileage_delta:.1f}"
    else:
        avg_mileage_delta_str = "N/A"
    
    summary_statistics = {
        "Average Weekly Mileage": f"{weekly_mileage_df['distance (miles)'].mean():.1f}",
        "Lowest Weekly Mileage": f"{weekly_mileage_df['distance (miles)'].min():.1f}",
        "Highest Weekly Mileage": f"{weekly_mileage_df['distance (miles)'].max():.1f}",
        "Average Change in Weekly Mileage": avg_mileage_delta_str,
    }
    return pd.DataFrame([summary_statistics])


def create_weekly_xtraining_df(all_activities: dict, group_by_sport_type: bool = False) -> pd.DataFrame:
    """
    Create a DataFrame with all cross training (i.e. non-running) activities.
    DataFrame will group training volume by week, and sport type if specified.

    Args:
        all_activities (dict): dictionary containing all Strava activities
        group_by_sport_type (bool, optional): Flag to group training volume by sport type. Defaults to False.

    Returns:
        pd.DataFrame: DataFrame containing weekly cross-training volume, optionally grouped by sport type
    """
    # Store cross training activities in a DataFrame
    xtraining_activities_list = []
    for activity_type, activity_list in all_activities.items():
        if activity_type != "run":
            xtraining_activities_list.extend(activity_list)
    df = pd.DataFrame(xtraining_activities_list)

    # Convert 'date' column to datetime
    df['start_date'] = pd.to_datetime(df['start_date'])

    # Group cross training volume by week and optionally by sport type
    if group_by_sport_type:
        weekly_totals = df.groupby(['sport_type', pd.Grouper(key='start_date', freq=WEEK_FREQ)])['elapsed_time'].sum().reset_index()
    else: 
        weekly_totals = df.groupby(pd.Grouper(key='start_date', freq=WEEK_FREQ))['elapsed_time'].sum().reset_index()

    # Format 'date' column and create a new column for year-week
    weekly_totals['date'] = weekly_totals['start_date'].dt.strftime("%Y-%m-%d")
    weekly_totals['year-week'] = weekly_totals['start_date'].dt.strftime(DATE_FORMAT)
    
    # Specify column order to write to the excel workbook
    first_cols = ['date', 'year-week']
    weekly_totals = weekly_totals[first_cols + [col for col in weekly_totals.columns if col not in first_cols]]

    return weekly_totals


def xtraining_vs_race_statistics(activities: dict) -> pd.DataFrame:
    """
    Compute summary statistics analyzing cross training activity vs. race times.
    Eventually, this may include statistics to quantify the relationship between cross-training volume and race performance.

    Args:
        activities (dict): dictionary of activities

    Returns:
        pd.DataFrame: DataFrame containing weekly cross-training statistics
    """
    xtraining_df = create_weekly_xtraining_df(activities)

    # Summary statistics should show elapsed time in hh:mm:ss format
    xtraining_df['elapsed_time'] = pd.to_timedelta(xtraining_df['elapsed_time'], unit='s')
    
    def convert_timedelta_to_hhmmss(td: timedelta) -> str:
        """
        Convert timedelta to a string that follows the format {hh:mm:ss}.
        I.e. the days should be omitted (ex: 1 day 00:00:00 -> 24:00:00)
        """
        # Calculate absolute pieces from total seconds
        total_seconds = int(td.total_seconds())
        hours = total_seconds // 3600
        minutes = (total_seconds % 3600) // 60
        seconds = total_seconds % 60

        # Format with 2-digit zero padding
        formatted_time = f"{hours:02d}:{minutes:02d}:{seconds:02d}"
        return formatted_time

    summary_statistics = {
        "Average Weekly X-Training Volume": convert_timedelta_to_hhmmss(xtraining_df['elapsed_time'].mean()),
        "Lowest Weekly X-Training Volume": convert_timedelta_to_hhmmss(xtraining_df['elapsed_time'].min()),
        "Highest Weekly X-Training Volume": convert_timedelta_to_hhmmss(xtraining_df['elapsed_time'].max())
    }
    return pd.DataFrame([summary_statistics])


################################
# PLOTTING - Visualize race performance against various factors like mileage, cross training, etc.
################################
def seconds_to_hhmmss(x, pos) -> str:
    """
    Format elapsed time measured in seconds into hh:mm:ss.

    Args:
        x (_type_): The raw numeric value of the tick mark (seconds)
        pos (_type_): The tick position (automatically handled by Matplotlib)

    Returns:
        str: _description_
    """
    # Handle negative values safely if your dataset has them
    is_negative = x < 0
    total_seconds = int(abs(x))
    
    # Calculate hours, minutes, and remaining seconds
    hours = total_seconds // 3600
    minutes = (total_seconds % 3600) // 60
    seconds = total_seconds % 60

    sign = "-" if is_negative else ""
    return f"{sign}{hours:02d}:{minutes:02d}:{seconds:02d}"


def visualize_weekly_mileage_vs_race(running_activities: dict, personal_records: dict, race: str):
    """
    Plot weekly mileage against race times.

    Args:
        running_activities (dict): all running activities
        personal_records (dict): all races/PRs
        race (str): race distance to analyze

    Returns:
        _type_: _description_
    """
    weekly_totals_df = create_weekly_mileage_df(running_activities)
    
    # 1. CREATE THE MATPLOTLIB FIGURE & SPECIFY THE REPORT LAYOUT
    # The Figure will contain two subplots: one for the weekly mileage plot, another for the table of summary statistics
    fig = plt.figure(
        figsize=FIG_SIZE,
        constrained_layout=True
    )
    gs = GridSpec(
        2, # 2 rows: weekly mileage plot & summary statistics
        1,
        figure=fig,
        height_ratios=[2, 1] # Weekly mileage plot will take up 2x as much space as the statistics table
    )


    # 2. CREATE THE WEEKLY MILEAGE PLOT
    # Place the plot at the top of the report
    ax1 = fig.add_subplot(gs[0])
    
    # Plot bar chart of weekly mileage
    ax1.bar(weekly_totals_df['year-week'], weekly_totals_df['distance (miles)'], color='darkturquoise', edgecolor='black') # Matplotlib colors: https://stackoverflow.com/questions/22408237/named-colors-in-matplotlib
    ax1.set_xlabel('Date (Year-Week)')

    # Set x-axis tick frequency to every 5 weeks to avoid crowding the plot 
    indices = range(0, len(weekly_totals_df), WEEKLY_MILEAGE_PLOT_TICK_FREQUENCY) # indices of dates to include
    ax1.set_xticks(indices)
    ax1.set_xticklabels(weekly_totals_df['year-week'].iloc[indices], rotation=90) 
    
    # Set the first y-axis for weekly mileage
    ax1.set_ylabel('Miles', color='darkturquoise')
    ax1.tick_params(axis='y', labelcolor='darkturquoise')

    # Create a secondary y-axis to plot race times against a shared x-axis (time)
    ax2 = ax1.twinx()

    # Create a line plot for race times
    pr_df = create_pr_df(personal_records, race)
    ax2.plot(pr_df['year-week'], pr_df['time (seconds)'], 'o--r', label='Race Time', color='gold')
    ax2.set_ylabel('Race Time (hh:mm:ss)', color='gold')
    ax2.tick_params(axis='y', labelcolor='gold')
    
    # Convert race time to mm:ss format
    ax2.yaxis.set_major_formatter(ticker.FuncFormatter(seconds_to_hhmmss))

    # Annotate the race times on the plot
    for _, row in pr_df.iterrows():
        ax2.annotate(
            f"({row['date']}, {seconds_to_hhmmss(row['time (seconds)'], None)})",
            (row['year-week'], row['time (seconds)']),
            xytext=(0, 5), # Offset: 0 points horizontally, 5 points vertically up
            textcoords='offset points', # Treat xytext as pixel offset
            fontsize=8,
            color='gold',
            weight='bold',
            ha="center", # Centers the text horizontally over the point
            va="bottom", # Positions the bottom of the text box above the point
        )
        
    # Add a legend & title
    handles1, labels1 = ax1.get_legend_handles_labels()
    handles2, labels2 = ax2.get_legend_handles_labels()
    ax1.legend(
        handles1 + handles2,
        labels1 + labels2,
        loc="lower center",
        bbox_to_anchor=(0.5, 1.1),
        ncol=4,
        frameon=True
    )

    plt.title(f'Weekly Mileage vs. Race Times ({race})')

    
    # 3. COMPUTE SUMMARY STATISTICS 
    # Place the summary statistics at the bottom of the report
    ax_table = fig.add_subplot(gs[1])
    ax_table.axis("off")
       
    summary_stats_df = mileage_vs_race_statistics(running_activities)
    table = ax_table.table(
        cellText=summary_stats_df.values,
        colLabels=summary_stats_df.columns,
        loc="center",
        cellLoc="center"
    )

    # Set the font size and bold the header row of the table
    table.auto_set_font_size(False)
    table.set_fontsize(10)
    table.scale(1.2, 2.0)
    for (row, col), cell in table.get_celld().items():
        if row == 0:
            cell.set_text_props(weight='bold')

    return fig, weekly_totals_df


def visualize_weekly_xtraining_vs_race(activities_dict: dict, personal_records: dict, race: str):
    """
    Visualize weekly cross training volume against race times.
    Cross training volume will be displayed as a bar chart, and race times will be plotted as a line chart.
    Each bar will display the distribution of that week's cross training volume broken down by activity (ex: swim, bike, lift, etc.)

    Args:
        activities_dict (dict): all activities
        personal_records (dict): all PRs/races
        race (str): race distance to analyze

    Returns:
        _type_: _description_
    """
    
    # 1. Create a pivot table containing cross training volume grouped by sport type & week
    weekly_xtraining_df = create_weekly_xtraining_df(activities_dict, group_by_sport_type=True)
    weekly_xtraining_df_pivoted = weekly_xtraining_df.pivot(index='year-week', columns='sport_type', values='elapsed_time')
    
    # 2. CREATE THE MATPLOTLIB FIGURE & SPECIFY THE REPORT LAYOUT
    # The Figure will contain two subplots: one for the weekly x-training plot, another for the table of summary statistics
    fig = plt.figure(
        figsize=FIG_SIZE,
        constrained_layout=True
    )
    gs = GridSpec(
        2, # 2 rows: weekly x-training plot & summary statistics
        1,
        figure=fig,
        height_ratios=[2, 1] # Weekly x-training plot will take up 2x as much space as the statistics table
    )
    
    # 3. PLOT WEEKLY X-TRAINING VOLUME
    ax1 = fig.add_subplot(gs[0]) 
    
    # Create the x-axis (time): determine the time range that captures all of the x-training activities and the races.
    # Then, plot the x-training activities & races onto their appropriate locations on the time axis.
    pr_df = create_pr_df(personal_records, race)
    
    # Convert week strings to dates
    xtraining_dates = pd.to_datetime(
        weekly_xtraining_df_pivoted.index + "-0", # need this day suffix to convert date string (%Y-W%U) to datetime (%Y-W%U-%w)
        format="%Y-W%U-%w"
    )
    pr_dates = pd.to_datetime(
        pr_df["year-week"] + "-0",
        format="%Y-W%U-%w"
    )
    
    # Find min & max weeks across all x-training & race activities, then generate every week in between
    start_week = min(
        xtraining_dates.min(),
        pr_dates.min()
    )
    end_week = max(
        xtraining_dates.max(),
        pr_dates.max()
    )
    all_weeks = pd.date_range(
        start=start_week,
        end=end_week,
        freq=WEEK_FREQ
    )
    
    # Convert weeks into their date strings
    week_labels = [
        d.strftime(DATE_FORMAT)
        for d in all_weeks
    ]
    
    # Map weeks to their xposition on the plot
    week_to_xpos = {
        week: i
        for i, week in enumerate(week_labels)
    }
    
    # Reindex the pivot table to include all weeks
    weekly_xtraining_df_pivoted = (
        weekly_xtraining_df_pivoted
        .reindex(week_labels)
        .fillna(0)
    )
    
    # Create the bar chart of weekly x-training volume
    weekly_xtraining_df_pivoted.plot(kind="bar", stacked=True, edgecolor="black", ax=ax1)
    ax1.set_xlabel('Date (Year-Week)')

    # Set the weekly frequency on the x-axis
    frequency = 4 # number of weeks between x ticks
    indices = range(0, len(weekly_xtraining_df_pivoted.index), frequency) # indices of dates to include
    ax1.set_xticks(indices)
    ax1.set_xticklabels(weekly_xtraining_df_pivoted.index[indices], rotation=90) 

    # Format y-axis (elapsed time) in terms of hours:minutes:seconds 
    ax1.set_ylabel('Elapsed Time (hh:mm:ss)', color='darkturquoise')
    ax1.tick_params(axis='y', labelcolor='darkturquoise')
    ax1.yaxis.set_major_formatter(ticker.FuncFormatter(seconds_to_hhmmss))

    # 4. PLOT RACE TIMES
    # Create a secondary y-axis (on the right of the plot) to plot race times against a shared x-axis (time)
    ax2 = ax1.twinx()

    # Map weeks in PR data to align with the index positions of the cross training dataframe.
    # When plotting the xtraining pivot table, pandas creates its own internal x-axis that maps integers to weeks (i.e. 0 -> 2024-W01, 1 -> 2024-02, etc.)
    # We need to map each week to its corresponding x-position on the plot to ensure the PRs are plotted at the correct weeks.
    pr_df['xpos'] = pr_df['year-week'].map(week_to_xpos)
    
    # Plot race times as dots attached by dashed line
    ax2.plot(pr_df['xpos'], pr_df['time (seconds)'], 'o--r', label='Race Time', color='gold') # plot points as circles connected by dashed line
    ax2.set_ylabel('Race Time (hh:mm:ss)', color='gold')
    ax2.tick_params(axis='y', labelcolor='gold')
    
    # Convert elapsed time to mm:ss format
    ax1.yaxis.set_major_formatter(ticker.FuncFormatter(seconds_to_hhmmss))
    ax2.yaxis.set_major_formatter(ticker.FuncFormatter(seconds_to_hhmmss))

    # Annotate the race times
    for _, row in pr_df.iterrows():
        ax2.annotate(
            f"({row['date']}, {seconds_to_hhmmss(row['time (seconds)'], None)})",
            (row['xpos'], row['time (seconds)']),
            xytext=(0, 5), # Offset: 0 points horizontally, 5 points vertically up
            textcoords='offset points', # Treat xytext as pixel offset
            fontsize=8,
            color='gold',
            weight='bold',
            ha="center", # Centers the text horizontally over the point
            va="bottom", # Positions the bottom of the text box above the point
        )
        
    # Add legend to distinguish sport types & race times
    handles1, labels1 = ax1.get_legend_handles_labels()
    handles2, labels2 = ax2.get_legend_handles_labels()
    ax1.legend(
        handles1 + handles2,
        labels1 + labels2,
        loc="lower center",
        bbox_to_anchor=(0.5, 1.1),
        ncol=4,
        frameon=True
    )

    # Add plot title
    plt.title(f'Weekly Cross Training Volume vs. Race Times ({race})')

    # 5. COMPUTE SUMMARY STATISTICS
    ax_table = fig.add_subplot(gs[1])
    ax_table.axis("off")
    
    summary_stats_df = xtraining_vs_race_statistics(activities_dict)
    table = ax_table.table(
        cellText=summary_stats_df.values,
        colLabels=summary_stats_df.columns,
        loc="center",
        cellLoc="center"
    )

    # Set the font size and bold the header row of the table
    table.auto_set_font_size(False)
    table.set_fontsize(10)
    table.scale(1.2, 2.0)
    for (row, col), cell in table.get_celld().items():
        if row == 0:
            cell.set_text_props(weight='bold')

    return fig, weekly_xtraining_df_pivoted



################################
# GENERAL UTILITIES
################################
def parse_date(date_str) -> datetime.date:
    """Define date string format required for input arguments.

    Args:
        date_str (_type_): _description_

    Returns:
        _type_: _description_
    """
    return datetime.strptime(date_str, "%Y-%m-%d").date()


def validate_input_arguments(args, parser):
    """
    Validate the input arguments specified by the user.
    
    Args:
        args (_type_): _description_
        parser (_type_): _description_
    """
    if args.start_date and args.end_date:
        if args.start_date > args.end_date:
            parser.error("Start date must be before the end date.")
            
        # Start & end dates should be at least a week apart to generate plots showing weekly mileage/x-training trends
        if (args.end_date - args.start_date) < timedelta(weeks=1):
            raise ValueError(f"Specified date range ({args.start_date} - {args.end_date}) is less than a week apart! Plots require at least a week of data.")
        
    
def write_dict_to_json(data: dict, output_dir: Path, file_name: str): 
    """
    Write the dictionary to a JSON file in the specified output directory.

    Args:
        data (dict): dictionary to write to JSON
        output_dir (Path): output directory path 
        file_name (str): file name to write to (excluding the extension)
    """
    filename = f"{file_name}.json"
    output_dir.mkdir(parents=True, exist_ok=True)
    output_file_path = output_dir / filename
    
    with open(output_file_path, "w") as f:
        json.dump(data, f)
    
    print(f"Successfully wrote to '{output_file_path}'.")


################################
# FILE OUTPUT
################################
def generate_pdf(filename: str, pdf_config: PdfConfig):
    """
    Create a PDF report with the specified filename.

    Args:
        filename (str): name of the PDF report
        pdf_config (PdfConfig): Class containing the figures to include in the PDF
    """
    filename = filename.strip().split(".")[0] + ".pdf" # if filename included extension, remove it and enforce .pdf
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    output_path = OUTPUT_DIR / filename
    
    with PdfPages(output_path) as pdf:
        for fig in pdf_config.figs:    
            pdf.savefig(fig)
            plt.close(fig)


def generate_xlsx(filename: str, xlsx_config: list[WorkbookSheetConfig]):
    """
    Create an excel workbook under the specified filename.

    Args:
        filename (str): name of the excel workbook to create
        xlsx_config (list[WorkbookSheetConfig]): Data to include in each worksheet in the excel workbook
    """
    filename_xlsx = filename.strip().split(".")[0] + ".xlsx" # if filename included extension, remove it and enforce .xlsx
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    output_path = OUTPUT_DIR / filename_xlsx 
    
    # Write data to the workbook: plots will be placed to the right of the dataframe, if applicable
    with pd.ExcelWriter(
        output_path,
        engine="openpyxl"
    ) as writer:
        
        XTRAINING_SHEET_NAME = "X-Training"
        
        # Create each worksheet in the excel workbook
        for xlsx_sheet in xlsx_config:
            # Excel doesn't accept timezone aware columns: we'll remove timezone info, but preserve local time representation
            xlsx_sheet.data = xlsx_sheet.data.apply(
                lambda col: col.dt.tz_localize(None)
                if isinstance(col.dtype, DatetimeTZDtype)
                else col
            )
            
            # For X-Training data, we'll create an extra header row to annotate all the sport type columns
            if xlsx_sheet.name == XTRAINING_SHEET_NAME:
                xlsx_sheet.data.to_excel(
                    writer, 
                    sheet_name=xlsx_sheet.name, 
                    index=True,
                    startrow=1 # Write data on the second row to make room for header row above sport-type columns
                )
                
                ws = writer.sheets[xlsx_sheet.name]
                
                # Get sport-type columns to merge
                first_data_col = 2  # column B -> Index occupies column A
                last_data_col = len(xlsx_sheet.data.columns) + 1

                # Merge cells above the sport type columns and add label: "Elapsed Time by Sport Type"
                ws.merge_cells(
                    start_row=1,
                    start_column=first_data_col,
                    end_row=1,
                    end_column=last_data_col
                )
                cell = ws.cell(row=1, column=first_data_col)
                cell.value = "Elapsed Time by Sport Type (seconds)"
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center")
                for cell in ws[2]:
                    cell.font = Font(bold=True)
            else:
                # For all other sheets, just write the data on the first row and bold the header
                xlsx_sheet.data.to_excel(
                    writer, 
                    sheet_name=xlsx_sheet.name, 
                    index=False
                )
                
                ws = writer.sheets[xlsx_sheet.name]
                
                for cell in ws[1]: # Header row is row 1 by default
                    cell.font = Font(bold=True)
            
            # If there's a figure to include, write it to the right of the data
            if xlsx_sheet.fig:
                # save figure to image (PNG)
                buffer = BytesIO()
                xlsx_sheet.fig.savefig(buffer, format="png", bbox_inches="tight") # svg for max zoom?
                buffer.seek(0)
                img = XLImage(buffer)
                
                # Write the image to the excel sheet: leave 2 blank columns between table and plot
                ncols = len(xlsx_sheet.data.columns)
                image_col = get_column_letter(ncols + 3)
                ws.add_image(img, f"{image_col}1") # write the image to the specific cell


def generate_reports(activities: dict, pr_map: dict, race: str, filename: str):
    """
    Generate a PDF report & Excel workbook containing the plots & underlying data.

    Args:
        activities (dict): all activities
        pr_map (dict): all races/PRs
        race (str): race distance to analyze
        filename (str): name of the PDF report & Excel workbook
    """
    # Create dataframes & figures
    mileage_vs_race_fig, mileage_vs_race_df = visualize_weekly_mileage_vs_race(activities["run"], pr_map, race)
    xtraining_vs_race_fig, xtraining_vs_race_df = visualize_weekly_xtraining_vs_race(activities, pr_map, race)
    pr_df = create_pr_df(pr_map, race)
    
    # Create PDF report
    generate_pdf(
        filename=filename, 
        pdf_config=PdfConfig(figs=[mileage_vs_race_fig, xtraining_vs_race_fig])
    )
    print(f"PDF report `{filename}.pdf` created!")
    
    # Create Excel workbook
    generate_xlsx(
        filename=filename,
        xlsx_config=[
            WorkbookSheetConfig(name="Mileage", data=mileage_vs_race_df, fig=mileage_vs_race_fig),
            WorkbookSheetConfig(name="X-Training", data=xtraining_vs_race_df, fig=xtraining_vs_race_fig),
            WorkbookSheetConfig(name=f"{race} PRs", data=pr_df)
        ]
    )
    print(f"Excel workbook `{filename}.xlsx` created!")
        

def main():
    # 1. Initialize the parser & parse CLI args
    parser = argparse.ArgumentParser(description="Strava Analyzer: visualize the relationship between your training & running performance!")
    parser.add_argument("-d", "--distance", type=str, required=True, choices=RACE_TYPES, help="The race distance you want to analyze (ex: 1 mile, 5K, 10K, Half Marathon, Marathon).")
    parser.add_argument("-n", "--num-races", type=int, default=5, help="The fastest N races you want to analyze for the distance (default: 5).")
    parser.add_argument("--start-date", type=parse_date, help="Start date (YYYY-MM-DD)")
    parser.add_argument("--end-date", type=parse_date, help="End date (YYYY-MM-DD)")
    parser.add_argument("--use-cached", action='store_true', help="Fetch activity & PR data from local JSON files instead of the Strava API (default: False).")
    parser.add_argument("-o", "--output", type=str, help="Filename of the analysis (default: {distance}_report_{timestamp}.pdf/.xlsx).")
    args = parser.parse_args()   
    
    # Validate the input arguments (such as date range)
    validate_input_arguments(args, parser)

    # Determine attributes needed to generate the reports (e.g. race type, output file name, number of races)
    RACE_TYPE = args.distance
    TOP_N = args.num_races 
    if args.output is None:
        day_timestamp = datetime.now().strftime("%Y%m%d")
        OUTPUT_FILE = f"{args.distance}_analysis_{day_timestamp}"
    else:
        OUTPUT_FILE = args.output
    
    
    if not args.use_cached:
        # Load data from the Strava API, if not operating in cached mode
        activities_map, pr_map = get_activities_and_prs(args.start_date, args.end_date, TOP_N)
    else:
        print(f"Loading activities & PRs from JSON files in '{DATA_DIR}'...")
        try:
            with open(f"{DATA_DIR}/{ACTIVITIES_FILE}.json") as f:
                activities_map = json.load(f)
            with open(f"{DATA_DIR}/{PR_FILE}.json") as f:
                pr_map = json.load(f)
            
            # Check if local files contain sufficient PR data.
            # If not, call the Strava API to fetch all activities & PRs.
            if not validate_pr_map(pr_map, TOP_N):
                print("Local PR data doesn't contain sufficient data, calling Strava API to fetch all activities.")
                activities_map, pr_map = get_activities_and_prs(args.start_date, args.end_date, TOP_N)
            
        except FileNotFoundError as e:
            print(f"Activity data not found from local JSON files: {e}.")
            print("Loading data from Strava API instead!")
            activities_map, pr_map = get_activities_and_prs(args.start_date, args.end_date, TOP_N)

    # Create reports (.pdf & .xlsx)
    generate_reports(activities_map, pr_map, RACE_TYPE, OUTPUT_FILE)
    

if __name__ == "__main__":
    main()
    